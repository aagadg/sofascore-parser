import openpyxl
from openpyxl import load_workbook
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
import re
from selenium.webdriver.support.ui import WebDriverWait

def findBlockHeight(soup):
    height = soup.find('div', attrs={'class': 'ReactVirtualized__Grid__innerScrollContainer'})
    try:
        b = re.split(': |;', str(height['style']))
        l = b[b.index(' max-height') + 1]
    except:
        return 0
    return round(float(l[:len(l)-2])*2)

def findMatchesArray(h,a):
    football=[]
    while (h <= a):
        driver.execute_script(f"window.scrollTo(0, {h});")
        WebDriverWait(driver, 3)
        data = driver.page_source
        soup = BeautifulSoup(data, "html.parser")
        h += 1200
        t = soup.find_all('a', attrs={'data-id': True})
        try:
            for i in t:
                tteams = i.find('div', attrs={"class": "sc-hLBbgP eIlfTT"})
                line = str()
                for elem in tteams:
                    # print(elem.text.strip())
                    line += elem.text.strip() + ":"
                score = i.find_all('div',
                                   attrs={'class': "sc-hLBbgP sc-eDvSVe fuUKnP bMwHQt sc-9199a964-2 kgwLqG score-box"})
                if score != []:
                    line += score[0].text[0] + '/' + score[1].text[0] + ":"
                live = i.find('div', attrs={'color': 'sofaSingles.live'})
                if live != None:
                    line += 'live'
                if line not in football:
                    football.append(line)
        except:
            z = 0
    return football

driver = webdriver.Chrome()
driver.get("https://www.sofascore.com/")
data = driver.page_source
soup = BeautifulSoup(data, "html.parser")
allmatches=[]
#=======Find-all-urls-of sports=========================================================
urls=[]
x=[]
p=soup.find('div',attrs={'class':'sc-hLBbgP dRtNhU sc-12472a74-0 ijBjmq'}).find_all('a')
for elem in p:
    uurl='https://www.sofascore.com'+elem['href']
    if uurl not in urls and uurl!='https://www.sofascore.com/motorsport':
        urls.append(uurl)
#=======end============================================================================
#========Iterating through urls =======================================================
for elem in urls:
    driver.get(elem)
    WebDriverWait(driver, 3)
    data = driver.page_source
    soup = BeautifulSoup(data, "html.parser")
#=======Find-matches-block-height=====================================================
    a = findBlockHeight(soup)
    h = 0
#======================================================================================
#========Iterating through page========================================================
    allmatches.append(findMatchesArray(h,a))
#========Creating new xlsx file========================================================
path = r'D:\Python prj\second\1.xlsx'
wb = openpyxl.Workbook()
wb.save(path)
#======================================================================================
#========opening created xlsx file and filling with dara===============================
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
for i in range(len(allmatches)):
    for j in range(len(allmatches[i])):
        x.append(allmatches[i][j].split(':'))
    name=urls[i].split('/')[-1]
    if name=='':
        name='football'
    df=pd.DataFrame(x)
    df.to_excel(writer,sheet_name=f'{name}',index=False, header=False)
    x=[]
#=====================================================================================
writer.close()
driver.close()
